import pandas as pd
import numpy as np
import pathlib 
import glob
import openpyxl as exel
import re
from datetime import datetime
from utils.logger import LoggerObj
import os

CASE_FIRST='○患者'
AGE_ONE='１ '
GENDER_TWO='２ '
PLACE_THREE='３ '
PROCESS_FOUR='４ '
MOVE_FIVE='５ '
CLASS_SIX='６ '
AFTER_SEVEN='７ '
SEPARATER='：'
inputDir=pathlib.Path('data')
inputFiles=inputDir.glob('*.txt')

patcaseno=re.compile('[１２３４５６７８９][０１２３４５６７８９][０１２３４５６７８９]')
patmmdd=re.compile('[１２]*[１２３４５６７８９]月[１２３]*[０１２３４５６７８９]日')
convertDict={
    '１':'1',
    '２':'2',
    '３':'3',
    '４':'4',
    '５':'5',
    '６':'6',
    '７':'7',
    '８':'8',
    '９':'9',
    '０':'0'
}
transTable=str.maketrans(convertDict)

# ファイル名:出力内容のリスト
outputData={}
logObj=LoggerObj('readpdf')
log=logObj.createLog('readpdf')

log.info('処理開始')
filenameList=[]
for fileName in inputFiles:
    filenameList.append(fileName.name)
    outputStrList={}
    outputMoveList={}
    outputProcessList={}
    log.info(fileName.name)
    fileDatas=[]
    isFirst=True
    with open(fileName,encoding='Shift-jis') as pfile:

        bufStr=''
        strList=[]
        isProcess=False
        isMove=False
        isMoveFirst=True
        moveText=[]
        processText=[]


        bfFirstTerm=''
        for textLine in pfile:
            tempSplitText=textLine.split(' ')
            firstTerm=tempSplitText[0]
            if isFirst:
                caseNo=patcaseno.search(firstTerm)
                strList.append(caseNo.group())
                bfFirstTerm=firstTerm
                isFirst=False
                continue
            if CASE_FIRST in firstTerm:
                fileDatas.append(strList)
                outputStrList[bfFirstTerm]=strList
                outputMoveList[bfFirstTerm]=moveText
                outputProcessList[bfFirstTerm]=processText
                bfFirstTerm=firstTerm
                strList=[]
                moveText=[]
                processText=[]
                caseNo=patcaseno.search(firstTerm)
                strList.append(caseNo.group())
                isMoveFirst=True
                continue

            if GENDER_TWO in textLine or PLACE_THREE in textLine \
                 or AGE_ONE in textLine or CLASS_SIX in textLine or AFTER_SEVEN in textLine:
                for text in tempSplitText:
                    if SEPARATER in text:
                        strList.append(text.split(SEPARATER)[1])
                continue

            if PROCESS_FOUR in textLine:
                isProcess=True
                isMove=False
            if MOVE_FIVE in textLine:
                isMove=True
                isProcess=False

            if isProcess:
                outputText=''.join(tempSplitText)
                inputDate=patmmdd.search(outputText)
                if  inputDate is not None:
                    # ここでヒットしないときの分岐を作る
                    otherText= patmmdd.split(outputText)[1]
                    tconvDate=inputDate.group()
                    tconvDate=tconvDate.translate(transTable)
                    tconvDate='2020年'+tconvDate
                    tconvDatetime = datetime.strptime(tconvDate, '%Y年%m月%d日')
                    convDate = tconvDatetime.strftime('%Y/%m/%d')
                    
                else:
                    if '４症状、経過：' in outputText:
                        continue
                    otherText=outputText
                    convDate=''
                output=convDate+otherText
                processText.append(output)
            if isMove:
                outputText=''.join(tempSplitText)
                if isMoveFirst:
                    isMoveFirst=False
                    moveText.append(outputText.split(SEPARATER)[1])
                    continue
                moveText.append(outputText)

        outputStrList[bfFirstTerm]=strList
        outputMoveList[bfFirstTerm]=moveText
        outputProcessList[bfFirstTerm]=processText
    outputData[fileName.name]=[outputStrList,outputMoveList,outputProcessList]

workbook=exel.load_workbook('template/output_template.xlsx')
sheet=workbook.active
sheet.title='ファイル出力'
rownum=3
date=datetime.now().strftime("%Y%m%d%H%M%S")
outputFile='outputFile'+'_'+date+'.xlsx'

for fileName in filenameList:
    datas=outputData[fileName]
    outputStrList=datas[0]
    outputMoveList=datas[1]
    outputProcessList=datas[2]

    for key,item in outputStrList.items():
        tex=outputProcessList[key]
        item.insert(3,''.join(tex))
        tex=outputMoveList[key]
        item.insert(4,''.join(tex))

        # 経過もここでマージ予定
        sheet.cell(column=1,row=rownum,value=rownum-2)
        sheet.cell(column=2,row=rownum,value=fileName)

        for i,text in enumerate(item):
            colnum=i+3
            sheet.cell(column=colnum,row=rownum,value=text)
        rownum=rownum+1
outputDir='output/'+date

os.makedirs(outputDir,exist_ok=True)

workbook.save( outputDir+'/'+outputFile)
log.info('処理終了')

